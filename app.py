import pandas as pd
import openpyxl
from io import BytesIO
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Function to create pivot tables
def create_pivot_tables(df):
    pivot_tables = {}

    # Column order for Slab pivot table
    column_order = ['>365', '>180', '>90', '>60', '<=60']

    # Pivot table 2: State_Count
    pivot_tables['State_Count'] = pd.pivot_table(df,
                                                 values='State_Count',
                                                 index=['State'],
                                                 columns='Slab',
                                                 aggfunc='sum',
                                                 fill_value=0)
    pivot_tables['State_Count'] = pivot_tables['State_Count'].reindex(columns=column_order, fill_value=0)
    pivot_tables['State_Count']['Grand Total'] = pivot_tables['State_Count'].sum(axis=1)
    pivot_tables['State_Count'] = pivot_tables['State_Count'].sort_values(by='Grand Total', ascending=False)
    grand_total_row = pivot_tables['State_Count'].sum(axis=0)
    grand_total_row.name = 'Total'
    pivot_tables['State_Count'] = pd.concat([pivot_tables['State_Count'], grand_total_row.to_frame().T])

    # Pivot table 1: State_Count by Slab
    pivot_tables['PDD Pending'] = pd.pivot_table(df,
                                                         values='State_Count',
                                                         index=['State', 'Cluster'],
                                                         columns='Slab',
                                                         aggfunc='sum',
                                                         fill_value=0)
    pivot_tables['PDD Pending'] = pivot_tables['PDD Pending'].reindex(columns=column_order, fill_value=0)
    pivot_tables['PDD Pending']['Grand Total'] = pivot_tables['PDD Pending'].sum(axis=1)
    grand_total_row = pivot_tables['PDD Pending'].sum(axis=0)
    grand_total_row.name = ('Total', 'Total')
    pivot_tables['PDD Pending'] = pd.concat([pivot_tables['PDD Pending'], grand_total_row.to_frame().T])

    # Pivot table 3: Total Discrepancy
    pivot_tables['Total Discrepancy'] = pd.pivot_table(df,
                                                       values=['Critical Count', 'Non Critical Count'],
                                                       index=['State', 'Cluster'],
                                                       aggfunc='sum',
                                                       fill_value=0)
    pivot_tables['Total Discrepancy']['Grand Total'] = pivot_tables['Total Discrepancy'].sum(axis=1)
    grand_total_row = pivot_tables['Total Discrepancy'].sum(axis=0)
    grand_total_row.name = ('Total', 'Total')
    pivot_tables['Total Discrepancy'] = pd.concat([pivot_tables['Total Discrepancy'], grand_total_row.to_frame().T])

    # Pivot table 4: Critical Count
    pivot_tables['Critical Count'] = pd.pivot_table(df,
                                                    values='Critical Count',
                                                    index=['State','Cluster'],
                                                    columns='Slab',
                                                    aggfunc='sum',
                                                    fill_value=0)
    pivot_tables['Critical Count'] = pivot_tables['Critical Count'].reindex(columns=column_order, fill_value=0)
    pivot_tables['Critical Count']['Grand Total'] = pivot_tables['Critical Count'].sum(axis=1)
    grand_total_row = pivot_tables['Critical Count'].sum(axis=0)
    grand_total_row.name = ('Total','Total')
    pivot_tables['Critical Count'] = pd.concat([pivot_tables['Critical Count'], grand_total_row.to_frame().T])

    # Pivot table 5: Non Critical Count
    pivot_tables['Non Critical Count'] = pd.pivot_table(df,
                                                        values='Non Critical Count',
                                                        index=['State', 'Cluster'],
                                                        columns='Slab',
                                                        aggfunc='sum',
                                                        fill_value=0)
    pivot_tables['Non Critical Count'] = pivot_tables['Non Critical Count'].reindex(columns=column_order, fill_value=0)
    pivot_tables['Non Critical Count']['Grand Total'] = pivot_tables['Non Critical Count'].sum(axis=1)
    grand_total_row = pivot_tables['Non Critical Count'].sum(axis=0)
    grand_total_row.name = ('Total', 'Total')
    pivot_tables['Non Critical Count'] = pd.concat([pivot_tables['Non Critical Count'], grand_total_row.to_frame().T])

    return pivot_tables

# Streamlit app
def main():
    st.title("PDD Pending")
    
    # File uploader
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
    
    if uploaded_file is not None:
        # Load the data from the uploaded file
        df = pd.read_excel(uploaded_file)

        # Ensure 'State' and 'Cluster' are properly formatted
        df['State'] = df['State'].astype(str).str.strip()
        df['Cluster'] = df['Cluster'].astype(str).str.strip()
        
        # Generate pivot tables
        pivot_tables = create_pivot_tables(df)
        
        # Define headers for each pivot table
        headers = {
            'PDD Pending': 'PDD Pending',
            'State_Count': 'PDD Pending',
            'Total Discrepancy': 'Total Discrepancy',
            'Critical Count': 'Critical Documents Pending',
            'Non Critical Count': 'Non Critical Documents Pending'
        }
        
        # Define border style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Save the pivot tables to a new Excel file
        output_file = BytesIO()
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, table in pivot_tables.items():
                # Rename "State" and "Cluster" headers for each pivot table
                if isinstance(table.index, pd.MultiIndex):
                    table.index.set_names(['State', 'Cluster'], inplace=True)
                else:
                    table.index.name = 'State'  # Renaming only 'State' if it's a single level index

                # Convert DataFrame to Excel sheet
                table.to_excel(writer, sheet_name=sheet_name, startrow=1, header=True)  # Adjusted startrow to 3
                worksheet = writer.sheets[sheet_name]

                # Add the custom main header with a merge and center
                header_text = headers[sheet_name]
                if sheet_name == 'State_Count':
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
                else:
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(table.columns) + 2)
                header_cell = worksheet.cell(row=1, column=1)
                header_cell.value = header_text
                header_cell.font = Font(bold=True, color="000000")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.fill = PatternFill(start_color="A7C6E6", end_color="A7C6E6", fill_type="solid")

                # Apply color to the existing second row (column headers) without adding extra rows
                header_row = 2  # This is the second row which contains your column headers
                num_columns = len(table.columns)  # Get the number of columns dynamically
                if sheet_name == 'State_Count':
                    columns_to_color = 7
                elif sheet_name=='Total Discrepancy':
                    columns_to_color=5  # Apply color to 7 columns in the 'State_Count' table
                else:
                    columns_to_color = 8  # Apply color to 8 columns in other pivot tables

                for col_idx in range(columns_to_color):
                    cell = worksheet.cell(row=header_row, column=col_idx + 1)  # 2nd row, adjusted column index

                    # Apply the background color (light green) to the header row
                    cell.fill = PatternFill(start_color="A7C6E6", end_color="A7C6E6", fill_type="solid")
                    cell.font = Font(bold=True,color="000000")  # Make the text bold
                    cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align text

                # Apply style to the Grand Total rows
                for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    if row[-1].value == 'Grand Total' or row[0].value == 'Total':  # Check if it's the total row
                        for cell in row:
                            cell.font = Font(bold=True, color="000000")  # Bold white text
                            cell.fill = PatternFill(start_color="A7C6E6", end_color="A7C6E6", fill_type="solid")  # Light blue background

                # Apply border to all cells
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border  # Apply the thin border style

        # Allow the user to download the processed file
        output_file.seek(0)
        st.download_button(label="Download Pivot Tables", data=output_file, file_name="pivot_tables.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    main()
